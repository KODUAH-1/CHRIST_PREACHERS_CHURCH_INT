from pathlib import Path
from app import create_app, db
from app.models import Branch, User, FundRecord, AttendanceRecord
from sqlalchemy import text
from openpyxl import load_workbook
from datetime import datetime, date
from decimal import Decimal
import re
import json


PROJECT_ROOT = Path(__file__).resolve().parent


print()
print("=" * 90)
print("CHRIST PREACHERS CHURCH - SAFE HISTORICAL IMPORT")
print("=" * 90)


# ============================================================
# 1. FIND EXCEL WORKBOOKS
# ============================================================

print()
print("=" * 90)
print("STEP 1 - WORKBOOK DISCOVERY")
print("=" * 90)

xlsx_files = sorted(
    PROJECT_ROOT.rglob("*.xlsx"),
    key=lambda p: p.stat().st_mtime,
    reverse=True
)

if not xlsx_files:
    print("NO XLSX FILES FOUND INSIDE PROJECT.")
    print()
    print("Place the combined workbook inside:")
    print(PROJECT_ROOT)
    raise SystemExit(1)

for i, path in enumerate(xlsx_files, 1):
    print(
        "{}. {} | {} bytes".format(
            i,
            path,
            path.stat().st_size
        )
    )

workbook_path = xlsx_files[0]

print()
print("SELECTED WORKBOOK:")
print(workbook_path)


# ============================================================
# 2. LOAD WORKBOOK
# ============================================================

print()
print("=" * 90)
print("STEP 2 - WORKBOOK SHEETS")
print("=" * 90)

wb = load_workbook(
    filename=workbook_path,
    data_only=True
)

for ws in wb.worksheets:
    print(
        "{} | rows={} | columns={}".format(
            repr(ws.title),
            ws.max_row,
            ws.max_column
        )
    )


# ============================================================
# 3. NORMALIZATION HELPERS
# ============================================================

def clean_text(value):
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).strip()
    )


def normalize_name(value):
    value = clean_text(value).upper()

    replacements = {
        "SAPEIMAN": "TAIFA",
        "SAPEIMAN ": "TAIFA",
        "N. NYANO": "NYANYANO",
        "N.NYANO": "NYANYANO",
        "N NYANO": "NYANYANO",
        "ODORKOR ": "ODORKOR",
        "KOFORID": "KONONGO",
        "NKAWKAW ": "NKAWKAW",
        "ASHAI": "ASHAIMAN",
        "ASHAIMAN ": "ASHAIMAN",
        "T.FA/MKS": "TAIFA",
    }

    return replacements.get(value, value)


def numeric(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text_value = clean_text(value)

    if not text_value:
        return None

    text_value = text_value.replace(",", "")

    try:
        return Decimal(text_value)
    except Exception:
        return None


def as_date(value):
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        value = value.strip()

        for fmt in (
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%m-%d-%Y",
        ):
            try:
                return datetime.strptime(value, fmt).date()
            except Exception:
                pass

    return None


# ============================================================
# 4. DATABASE SAFETY CHECK
# ============================================================

app = create_app()

with app.app_context():

    print()
    print("=" * 90)
    print("STEP 3 - DATABASE SAFETY CHECK")
    print("=" * 90)

    fund_count = db.session.execute(
        text("SELECT COUNT(*) FROM fund_records")
    ).scalar()

    attendance_count = db.session.execute(
        text("SELECT COUNT(*) FROM attendance_records")
    ).scalar()

    print("fund_records =", fund_count)
    print("attendance_records =", attendance_count)

    if fund_count != 0 or attendance_count != 0:
        print()
        print("IMPORT ABORTED.")
        print("Target tables are no longer empty.")
        print("No database changes were made.")
        raise SystemExit(2)

    print()
    print("SAFE: target historical tables are empty.")


    # ========================================================
    # 5. BRANCH MAP
    # ========================================================

    print()
    print("=" * 90)
    print("STEP 4 - CANONICAL BRANCH MAP")
    print("=" * 90)

    branches = Branch.query.order_by(Branch.id).all()

    branch_map = {}

    for branch in branches:
        canonical = normalize_name(branch.name)
        code = normalize_name(branch.code)

        branch_map[canonical] = branch
        branch_map[code] = branch

        print(
            "ID={} | NAME={} | CODE={} | NORMALIZED={}".format(
                branch.id,
                branch.name,
                branch.code,
                canonical
            )
        )


    # ========================================================
    # 6. SHEET DISCOVERY
    # ========================================================

    print()
    print("=" * 90)
    print("STEP 5 - SOURCE SHEET CLASSIFICATION")
    print("=" * 90)

    sheet_map = {}

    for ws in wb.worksheets:

        original = ws.title
        normalized = clean_text(original).upper()

        classification = "OTHER"

        if "STATEMENT OF ACCOUNT" in normalized:
            classification = "STATEMENT"

        elif "PAYMENT SCHEDULE" in normalized:
            classification = "PAYMENT_SCHEDULE"

        elif "NEW COMM" in normalized:
            classification = "NEW_COMERS"

        elif "HUBTEL" in normalized:
            classification = "HUBTEL"

        elif "SOWUTUOM" in normalized:
            classification = "SOWUTUOM"

        elif "BUILDING" in normalized:
            classification = "BUILDING_PROJECT"

        elif "ATTND" in normalized or "ATTEND" in normalized:
            classification = "ATTENDANCE"

        elif "FUND" in normalized:
            classification = "FUND"

        sheet_map[original] = classification

        print(
            "{} => {}".format(
                repr(original),
                classification
            )
        )


    # ========================================================
    # 7. RAW SOURCE INVENTORY
    # ========================================================

    print()
    print("=" * 90)
    print("STEP 6 - SOURCE INVENTORY")
    print("=" * 90)

    inventory = {
        "workbook": str(workbook_path),
        "sheets": {},
        "fund_candidates": 0,
        "attendance_candidates": 0,
        "unmapped_branches": [],
        "warnings": []
    }


    for ws in wb.worksheets:

        classification = sheet_map[ws.title]

        non_empty = 0

        for row in ws.iter_rows():

            values = [
                cell.value
                for cell in row
            ]

            if any(
                value is not None and clean_text(value)
                for value in values
            ):
                non_empty += 1

        inventory["sheets"][ws.title] = {
            "classification": classification,
            "rows": ws.max_row,
            "columns": ws.max_column,
            "non_empty_rows": non_empty
        }

        print(
            "{} | {} | non-empty rows={}".format(
                repr(ws.title),
                classification,
                non_empty
            )
        )


    # ========================================================
    # 8. ATTENDANCE DATE DETECTION
    # ========================================================

    print()
    print("=" * 90)
    print("STEP 7 - ATTENDANCE DATE DETECTION")
    print("=" * 90)

    attendance_dates = []

    for ws in wb.worksheets:

        classification = sheet_map[ws.title]

        if classification != "ATTENDANCE":
            continue

        for row in ws.iter_rows():

            for cell in row:

                value = cell.value

                parsed = as_date(value)

                if parsed:
                    attendance_dates.append(
                        (
                            ws.title,
                            cell.row,
                            cell.column,
                            parsed
                        )
                    )

    attendance_dates = sorted(
        set(attendance_dates),
        key=lambda x: (
            x[0],
            x[3],
            x[1],
            x[2]
        )
    )

    for item in attendance_dates:
        print(
            "SHEET={} | ROW={} | COL={} | DATE={}".format(
                item[0],
                item[1],
                item[2],
                item[3]
            )
        )

    print(
        "DETECTED ATTENDANCE DATE CELLS =",
        len(attendance_dates)
    )


    # ========================================================
    # 9. FUND SOURCE DETECTION
    # ========================================================

    print()
    print("=" * 90)
    print("STEP 8 - FUND SOURCE DETECTION")
    print("=" * 90)

    fund_keywords = {
        "national": [
            "NATIONAL",
            "N. FUND",
            "N FUND",
            "NATIONAL FUND"
        ],
        "evangelical": [
            "EVANGELICAL",
            "EVANGELICAL FUND",
            "E. FUND"
        ]
    }

    fund_keyword_hits = {
        "national": [],
        "evangelical": []
    }

    for ws in wb.worksheets:

        classification = sheet_map[ws.title]

        if classification not in (
            "FUND",
            "STATEMENT",
            "PAYMENT_SCHEDULE"
        ):
            continue

        for row in ws.iter_rows():

            for cell in row:

                text_value = clean_text(cell.value).upper()

                if not text_value:
                    continue

                for fund_type, keywords in fund_keywords.items():

                    for keyword in keywords:

                        if keyword in text_value:

                            fund_keyword_hits[fund_type].append(
                                (
                                    ws.title,
                                    cell.row,
                                    cell.column,
                                    text_value
                                )
                            )

                            break

    for fund_type, hits in fund_keyword_hits.items():

        print()
        print(
            "{} FUND KEYWORD HITS = {}".format(
                fund_type.upper(),
                len(hits)
            )
        )

        for hit in hits[:100]:
            print(
                "SHEET={} | ROW={} | COL={} | VALUE={}".format(
                    hit[0],
                    hit[1],
                    hit[2],
                    hit[3]
                )
            )


    # ========================================================
    # 10. BUILDING PROJECT INVENTORY
    # ========================================================

    print()
    print("=" * 90)
    print("STEP 9 - BUILDING PROJECT INVENTORY")
    print("=" * 90)

    building_rows = []

    for ws in wb.worksheets:

        if sheet_map[ws.title] != "BUILDING_PROJECT":
            continue

        for row_number, row in enumerate(
            ws.iter_rows(),
            start=1
        ):

            values = [
                cell.value
                for cell in row
            ]

            if not any(
                value is not None and clean_text(value)
                for value in values
            ):
                continue

            parsed_date = None
            name = None
            amount = None
            detail = None

            for value in values:

                if parsed_date is None:
                    parsed_date = as_date(value)

                if name is None:
                    text_value = clean_text(value)

                    if (
                        text_value
                        and not text_value.upper()
                        in (
                            "DATE",
                            "NAME",
                            "AMOUNT",
                            "DETAIL",
                            "CHURCH BUILDING"
                        )
                        and numeric(value) is None
                    ):
                        name = text_value

                if amount is None:
                    candidate = numeric(value)

                    if candidate is not None:
                        amount = candidate

                if (
                    detail is None
                    and isinstance(value, str)
                    and "BUILDING" in value.upper()
                ):
                    detail = clean_text(value)

            if parsed_date and name and amount is not None:

                building_rows.append(
                    {
                        "sheet": ws.title,
                        "row": row_number,
                        "date": str(parsed_date),
                        "name": name,
                        "amount": str(amount),
                        "detail": detail
                    }
                )

    print(
        "BUILDING PROJECT CANDIDATES =",
        len(building_rows)
    )

    for item in building_rows[:50]:
        print(
            "{} | {} | {} | {} | {}".format(
                item["date"],
                item["name"],
                item["amount"],
                item["detail"],
                item["sheet"]
            )
        )


    # ========================================================
    # 11. NEW COMERS INVENTORY
    # ========================================================

    print()
    print("=" * 90)
    print("STEP 10 - NEW COMERS INVENTORY")
    print("=" * 90)

    new_comers = []

    for ws in wb.worksheets:

        if sheet_map[ws.title] != "NEW_COMERS":
            continue

        for row_number, row in enumerate(
            ws.iter_rows(),
            start=1
        ):

            values = [
                cell.value
                for cell in row
            ]

            if row_number == 1:
                continue

            if not values:
                continue

            name = clean_text(
                values[0] if len(values) > 0 else ""
            )

            number = clean_text(
                values[1] if len(values) > 1 else ""
            )

            joined_date = (
                as_date(values[2])
                if len(values) > 2
                else None
            )

            if name:

                new_comers.append(
                    {
                        "sheet": ws.title,
                        "row": row_number,
                        "name": name,
                        "number": number,
                        "date": (
                            str(joined_date)
                            if joined_date
                            else None
                        )
                    }
                )

    print(
        "NEW COMER CANDIDATES =",
        len(new_comers)
    )

    for item in new_comers:
        print(
            "{} | {} | {} | {}".format(
                item["name"],
                item["number"],
                item["date"],
                item["sheet"]
            )
        )


    # ========================================================
    # 12. SUMMARY
    # ========================================================

    print()
    print("=" * 90)
    print("DRY-RUN IMPORT SUMMARY")
    print("=" * 90)

    print("Workbook:", workbook_path)
    print("Sheets:", len(wb.sheetnames))
    print("Attendance date cells:", len(attendance_dates))
    print("Building project candidates:", len(building_rows))
    print("New comer candidates:", len(new_comers))
    print("Fund target rows:", fund_count)
    print("Attendance target rows:", attendance_count)

    print()
    print("DATABASE CHANGES: 0")
    print()
    print("=" * 90)
    print("DRY RUN COMPLETE")
    print("=" * 90)


    # ========================================================
    # 13. SAVE MAPPING REPORT
    # ========================================================

    report_path = PROJECT_ROOT / "historical_import_mapping_report.json"

    with open(
        report_path,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            inventory,
            f,
            indent=2,
            ensure_ascii=False
        )

    print()
    print("Mapping report written to:")
    print(report_path)

